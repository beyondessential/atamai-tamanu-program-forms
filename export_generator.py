"""
Generates a human-readable questions export Excel from a ProgramDefinition.

This is NOT a Tamanu import file — it is a review document showing question text,
types, options, mandatory flags, and visibility conditions in a readable format.
One sheet per survey.
"""

import json
from io import BytesIO

import openpyxl
from openpyxl.styles import Font

from baml_client.types import ProgramDefinition


def _mandatory(validation_criteria: str | None) -> str:
    if not validation_criteria:
        return ""
    try:
        vc = json.loads(validation_criteria)
        return "Yes" if vc.get("mandatory") else ""
    except (json.JSONDecodeError, TypeError):
        return ""


def _visible_when(visibility_criteria: str | None) -> str:
    if not visibility_criteria:
        return ""
    try:
        vc = json.loads(visibility_criteria)
        conditions = vc.get("conditions", [])
        parts = []
        for c in conditions:
            qid = c.get("questionId", "")
            val = c.get("_value", "")
            cmp = c.get("_comparison", "=")
            parts.append(f"{qid} {cmp} {val}")
        join = " AND " if vc.get("_conjunction") == "and" else " OR "
        return join.join(parts)
    except (json.JSONDecodeError, TypeError):
        return visibility_criteria or ""


def generate_questions_export(program: ProgramDefinition) -> bytes:
    """Generate a human-readable review Excel from a ProgramDefinition (BAML type)."""
    wb = openpyxl.Workbook()
    first = True

    sheets_by_name = {s.survey_name: s for s in program.survey_sheets}

    for survey in program.surveys:
        sheet_data = sheets_by_name.get(survey.name)
        if not sheet_data:
            continue

        if first:
            ws = wb.active
            ws.title = survey.name[:31]
            first = False
        else:
            ws = wb.create_sheet(title=survey.name[:31])

        headers = ["Code", "Type", "Question Text", "Detail", "Options", "Mandatory", "Visible When", "Calculation"]
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True)

        for q in sheet_data.questions:
            q_type = q.type.value if hasattr(q.type, "value") else str(q.type)
            ws.append([
                q.code,
                q_type,
                q.text,
                q.detail or "",
                q.options or "",
                _mandatory(q.validation_criteria),
                _visible_when(q.visibility_criteria),
                q.calculation or "",
            ])

        for col in ws.columns:
            max_len = max((len(str(cell.value or "")) for cell in col), default=0)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 60)

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
