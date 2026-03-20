"""
Parses a Tamanu program importer XLSX into a text summary for the AI assistant,
a structured program_data dict for round-trip generation, and validation errors.
"""

import json
from io import BytesIO

import openpyxl

VALID_SURVEY_TYPES = {
    "programs", "vitals", "referral", "simpleChart",
    "complexChart", "complexChartCore", "obsolete",
}
VALID_STATUSES = {"publish", "draft", "hidden"}
VALID_VISIBILITY_STATUSES = {"current", "historical", "merged"}


def _cell_str(value) -> str:
    """Convert a cell value to a normalised string."""
    if value is None:
        return ""
    if isinstance(value, bool):
        return str(value).lower()  # True → "true", False → "false"
    return str(value).strip()


def _parse_kv_rows(ws) -> tuple[dict[str, str], int | None]:
    """
    Read key-value rows from the top of a sheet until a row whose first cell
    is 'code' or 'name' (the table header). Returns (kv_dict, header_row_number).
    header_row_number is 1-based; None if no header row was found.
    """
    kv: dict[str, str] = {}
    for row in ws.iter_rows():
        first = _cell_str(row[0].value) if row else ""
        if first.lower() in ("code", "name"):
            return kv, row[0].row
        if first and len(row) >= 2:
            kv[first] = _cell_str(row[1].value)
    return kv, None


def _parse_table(ws, header_row: int) -> list[dict[str, str]]:
    """Parse a table starting at header_row (1-based). Returns list of row dicts."""
    all_rows = list(ws.iter_rows(min_row=header_row))
    if not all_rows:
        return []
    headers = [_cell_str(c.value) for c in all_rows[0]]
    result = []
    for row in all_rows[1:]:
        values = [_cell_str(c.value) for c in row]
        row_dict = dict(zip(headers, values))
        if any(v for v in row_dict.values()):  # skip blank rows
            result.append(row_dict)
    return result


def _is_valid_json(value: str) -> bool:
    if not value:
        return True
    try:
        json.loads(value)
        return True
    except (json.JSONDecodeError, ValueError):
        return False


def parse_xlsx(data: bytes) -> tuple[str, dict | None, list[str]]:
    """
    Parse a Tamanu program importer XLSX.

    Returns (summary, program_data, errors):
    - summary: text for the AI conversation context; empty string on failure
    - program_data: structured dict for round-trip generation; None on failure
    - errors: list of validation error strings
    """
    errors: list[str] = []
    lines: list[str] = []

    try:
        wb = openpyxl.load_workbook(BytesIO(data), data_only=True)
    except Exception as e:
        return "", None, [f"Could not open file: {e}"]

    # ── Metadata ──────────────────────────────────────────────────────────────

    if "Metadata" not in wb.sheetnames:
        return "", None, ["Missing required sheet: Metadata"]

    ws_meta = wb["Metadata"]
    kv, header_row = _parse_kv_rows(ws_meta)

    program_name = kv.get("programName", "")
    program_code = kv.get("programCode", "")
    country = kv.get("country", "")

    if not program_name:
        errors.append("Metadata: programName is missing")
    if not program_code:
        errors.append("Metadata: programCode is missing")

    country_str = f", Country: {country}" if country else ""
    lines.append("[EXISTING PROGRAM LOADED]")
    lines.append(f"Program: {program_name or '(unnamed)'} ({program_code}){country_str}")
    lines.append("")

    surveys: list[dict] = []
    if header_row:
        survey_rows = _parse_table(ws_meta, header_row)
        if survey_rows:
            missing_cols = {"code", "name", "surveyType"} - set(survey_rows[0].keys())
            if missing_cols:
                errors.append(f"Metadata survey table missing columns: {', '.join(missing_cols)}")
        for s in survey_rows:
            s_name = s.get("name", "")
            s_type = s.get("surveyType", "")
            s_status = s.get("status", "draft")
            s_vis = s.get("visibilityStatus", "current")
            if s_type and s_type not in VALID_SURVEY_TYPES:
                errors.append(f"Survey '{s_name}': invalid surveyType '{s_type}'")
            if s_status and s_status not in VALID_STATUSES:
                errors.append(f"Survey '{s_name}': invalid status '{s_status}'")
            if s_vis and s_vis not in VALID_VISIBILITY_STATUSES:
                errors.append(f"Survey '{s_name}': invalid visibilityStatus '{s_vis}'")
            surveys.append(s)

    # ── Survey sheets ──────────────────────────────────────────────────────────

    program_surveys: list[dict] = []

    lines.append("Surveys:")
    for s in surveys:
        s_name = s.get("name", "")
        s_code = s.get("code", "")
        s_type = s.get("surveyType", "")
        s_status = s.get("status", "draft")
        sensitive = _cell_str(s.get("isSensitive", ""))
        notifiable = _cell_str(s.get("notifiable", ""))
        notify_emails = s.get("notifyEmailAddresses", "")

        meta_parts = [f"type: {s_type}", f"status: {s_status}"]
        if sensitive in ("true", "yes"):
            meta_parts.append("sensitive: yes")
        if notifiable in ("true", "yes"):
            emails = f" (emails: {notify_emails})" if notify_emails else ""
            meta_parts.append(f"notifiable: yes{emails}")

        lines.append(f"\n{s_name} ({s_code}) — {', '.join(meta_parts)}")

        # Mirror Tamanu's importSurvey.js lookup:
        # workbook.Sheets[sheetName.replace(/['"]/g, '')] || workbook.Sheets[code]
        s_name_clean = s_name.replace("'", "").replace('"', "")
        if s_name_clean in wb.sheetnames:
            ws_survey = wb[s_name_clean]
        elif s_code and s_code in wb.sheetnames:
            ws_survey = wb[s_code]
        else:
            errors.append(f"Missing sheet for survey '{s_name}'")
            lines.append("  (sheet not found)")
            program_surveys.append({**s, "questions": []})
            continue

        all_rows = list(ws_survey.iter_rows(min_row=1))
        if not all_rows:
            lines.append("  (empty sheet)")
            program_surveys.append({**s, "questions": []})
            continue

        headers = [_cell_str(c.value) for c in all_rows[0]]
        questions = []
        for row in all_rows[1:]:
            values = [_cell_str(c.value) for c in row]
            q = dict(zip(headers, values))
            if not q.get("code"):
                continue
            for field in ("visibilityCriteria", "validationCriteria", "config"):
                val = q.get(field, "")
                if val and not _is_valid_json(val):
                    errors.append(
                        f"Survey '{s_name}', question '{q['code']}': {field} is not valid JSON"
                    )
            vis = q.get("visibilityStatus", "")
            if vis and vis not in VALID_VISIBILITY_STATUSES:
                errors.append(
                    f"Survey '{s_name}', question '{q['code']}': invalid visibilityStatus '{vis}'"
                )
            questions.append(q)

        lines.append(f"  Questions ({len(questions)}):")
        for q in questions:
            desc = f"  - {q.get('code', '')}: {q.get('type', '')}"
            text = q.get("text", "") or q.get("name", "")
            if text:
                desc += f" — {text}"
            if q.get("newScreen") == "yes":
                desc += " [new screen]"
            if q.get("options"):
                desc += f" [options: {q['options']}]"
            if q.get("optionLabels"):
                desc += f" [labels: {q['optionLabels']}]"
            if q.get("detail"):
                desc += f" [detail: {q['detail']}]"
            if q.get("validationCriteria"):
                desc += f" [validation: {q['validationCriteria']}]"
            if q.get("visibilityCriteria"):
                desc += f" [visible when: {q['visibilityCriteria']}]"
            if q.get("calculation"):
                desc += f" [calculation: {q['calculation']}]"
            if q.get("config"):
                desc += f" [config: {q['config']}]"
            if q.get("indicator"):
                desc += f" [indicator: {q['indicator']}]"
            if q.get("visibilityStatus") and q.get("visibilityStatus") != "current":
                desc += f" [visibilityStatus: {q['visibilityStatus']}]"
            if q.get("visualisationConfig"):
                desc += f" [visualisationConfig: {q['visualisationConfig']}]"
            lines.append(desc)

        program_surveys.append({**s, "questions": questions})

    # ── Registry ───────────────────────────────────────────────────────────────

    registry_data: dict | None = None

    if "Registry" in wb.sheetnames:
        ws_reg = wb["Registry"]
        reg_kv, reg_header_row = _parse_kv_rows(ws_reg)
        reg_code = reg_kv.get("registryCode", "")
        reg_name = reg_kv.get("registryName", "")
        currently_at = reg_kv.get("currentlyAtType", "")
        lines.append(f"\nRegistry: {reg_name} ({reg_code}) — tracked at: {currently_at}")

        statuses = []
        if reg_header_row:
            statuses = _parse_table(ws_reg, reg_header_row)
            if statuses:
                status_list = ", ".join(
                    f"{s.get('name', '')} ({s.get('color', '')})" for s in statuses
                )
                lines.append(f"  Clinical statuses: {status_list}")

        conditions = []
        if "Registry Conditions" in wb.sheetnames:
            ws_cond = wb["Registry Conditions"]
            _, cond_header = _parse_kv_rows(ws_cond)
            if cond_header:
                conditions = _parse_table(ws_cond, cond_header)
                if conditions:
                    cond_list = ", ".join(r.get("name", "") for r in conditions if r.get("name"))
                    lines.append(f"  Conditions: {cond_list}")

        condition_categories = []
        if "Registry Condition Categories" in wb.sheetnames:
            ws_cat = wb["Registry Condition Categories"]
            _, cat_header = _parse_kv_rows(ws_cat)
            if cat_header:
                condition_categories = _parse_table(ws_cat, cat_header)
                if condition_categories:
                    cat_list = ", ".join(r.get("name", "") for r in condition_categories if r.get("name"))
                    lines.append(f"  Condition categories: {cat_list}")

        registry_data = {
            "kv": reg_kv,
            "statuses": statuses,
            "conditions": conditions,
            "condition_categories": condition_categories,
        }

    elif "Registry Conditions" in wb.sheetnames:
        ws_cond = wb["Registry Conditions"]
        _, cond_header = _parse_kv_rows(ws_cond)
        conditions = []
        if cond_header:
            conditions = _parse_table(ws_cond, cond_header)
            if conditions:
                cond_list = ", ".join(r.get("name", "") for r in conditions if r.get("name"))
                lines.append(f"\nRegistry Conditions: {cond_list}")

    program_data: dict = {
        "program_code": program_code,
        "program_name": program_name,
        "country": country,
        "surveys": program_surveys,
        "registry": registry_data,
    }

    return "\n".join(lines), program_data, errors
