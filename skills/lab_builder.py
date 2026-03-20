"""Lab Builder skill — creates Tamanu lab test reference data XLSX files.

Generates an XLSX compatible with Tamanu's referenceDataImporter. The file
contains up to four sheets:

  Lab Test Categories      — ReferenceData rows with type=labTestCategory
  Lab Test Types           — LabTestType rows with ranges as "min, max" strings
  Lab Test Panels          — LabTestPanel rows (optional)
  Lab Test Panel Lab Test Types — panel↔test links (optional)
"""

import re
from io import BytesIO
from types import SimpleNamespace

import openpyxl
import streamlit as st
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from baml_client.sync_client import b

TITLE = "Lab Builder"
ICON = "🧪"
DESCRIPTION = "Build Tamanu lab test reference data — categories, tests, and reference ranges."

_MANUAL = """
## What can this tool do?

Atamai helps you build Tamanu lab test reference data through conversation.
Describe the tests and their reference ranges — the AI asks clarifying questions then
generates a structured Excel file you can import into Tamanu.

---

### What it generates

An Excel file with sheets that match Tamanu's reference data importer:

**Lab Test Categories sheet** — one row per category:
- Code and name (e.g. "haematology" / "Haematology")

**Lab Test Types sheet** — one row per test:
- Code and name
- Category it belongs to
- Result type: Number, FreeText, or Select
- Unit of measurement (for Number tests, e.g. g/dL, mmol/L)
- Male and female reference ranges (min and max)
- Options (for Select tests)
- Sensitive flag

**Lab Test Panels sheet** *(optional)* — one row per panel:
- Code, name, and category

**Lab Test Panel Lab Test Types sheet** *(optional)* — links panels to tests

---

### Notes on reference ranges

Tamanu supports male and female reference ranges only — there is no age banding
for lab test types. If your ranges are the same for all sexes, the AI will set
both male and female ranges to the same values.

---

### Example prompt

> "I need reference ranges for a haematology panel.
> Include Haemoglobin, WBC, Platelets, and Haematocrit.
> Haemoglobin ranges differ by sex:
>   Male: 13.5–17.5 g/dL
>   Female: 12.0–15.5 g/dL
> WBC: 4.0–11.0 x10^9/L for all.
> Platelets: 150–400 x10^9/L for all.
> Haematocrit — male: 41–53%, female: 36–46%."

---

### Tips

- Each test must belong to a category — the AI will ask you to confirm
- Mention "panels" if you want grouped test ordering (e.g. a Full Blood Count panel)
- Result type defaults to Number; say "free text" or "select with options" to change it
"""


@st.dialog("Help & Manual", width="large")
def _show_manual() -> None:
    st.markdown(_MANUAL)


_HEADER_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")


def _fmt_range(min_val: float | None, max_val: float | None) -> str:
    """Format a min/max pair as Tamanu's 'N, N' range string, or empty string."""
    if min_val is None or max_val is None:
        return ""
    # Format without unnecessary trailing zeros
    def _n(v: float) -> str:
        return f"{v:g}"
    return f"{_n(min_val)}, {_n(max_val)}"


def _category_id(code: str) -> str:
    return f"labTestCategory-{code}"


def _test_type_id(code: str) -> str:
    return f"labTestType-{code}"


def _panel_id(code: str) -> str:
    return f"labTestPanel-{code}"


def _add_header(ws: object, headers: list[str]) -> None:
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = _HEADER_FILL


def _generate_xlsx(lab: object) -> bytes:
    """Convert a LabDefinition into a Tamanu referenceDataImporter-compatible XLSX."""
    wb = openpyxl.Workbook()

    # ── Lab Test Categories ─────────────────────────────────────────────────
    ws_cat = wb.active
    ws_cat.title = "Lab Test Categories"
    _add_header(ws_cat, ["id", "type", "code", "name", "visibilityStatus"])
    for cat in lab.categories:
        ws_cat.append([
            _category_id(cat.code),
            "labTestCategory",
            cat.code,
            cat.name,
            "current",
        ])
    for i, width in enumerate([30, 20, 20, 30, 15], 1):
        ws_cat.column_dimensions[get_column_letter(i)].width = width
    ws_cat.freeze_panes = "A2"

    # ── Lab Test Types ──────────────────────────────────────────────────────
    ws_tests = wb.create_sheet("Lab Test Types")
    _add_header(ws_tests, [
        "id", "code", "name", "labTestCategoryId",
        "resultType", "unit", "maleRange", "femaleRange",
        "options", "visibilityStatus", "isSensitive",
    ])
    for test in lab.tests:
        male_range = _fmt_range(test.male_min, test.male_max)
        female_range = _fmt_range(test.female_min, test.female_max)
        ws_tests.append([
            _test_type_id(test.code),
            test.code,
            test.name,
            _category_id(test.category_code),
            test.result_type,
            test.unit or "",
            male_range,
            female_range,
            test.options or "",
            "current",
            "TRUE" if test.is_sensitive else "FALSE",
        ])
    for i, width in enumerate([30, 20, 30, 30, 12, 12, 16, 16, 30, 15, 12], 1):
        ws_tests.column_dimensions[get_column_letter(i)].width = width
    ws_tests.freeze_panes = "A2"

    # ── Lab Test Panels (optional) ──────────────────────────────────────────
    # The referenceDataImporter uses labTestPanelLoader which reads a
    # testTypesInPanel column (comma-separated labTestTypeIds) and expands it
    # into LabTestPanelLabTestTypes rows internally — no separate sheet needed.
    if lab.panels:
        # Build panel_code → [labTestTypeId, ...] lookup from panel_test_links
        panel_tests: dict[str, list[str]] = {}
        for link in (lab.panel_test_links or []):
            panel_tests.setdefault(link.panel_code, []).append(_test_type_id(link.test_code))

        ws_panels = wb.create_sheet("Lab Test Panels")
        _add_header(ws_panels, ["id", "code", "name", "categoryId", "visibilityStatus", "testTypesInPanel"])
        for panel in lab.panels:
            test_ids = panel_tests.get(panel.code, [])
            ws_panels.append([
                _panel_id(panel.code),
                panel.code,
                panel.name,
                _category_id(panel.category_code),
                "current",
                ", ".join(test_ids),
            ])
        for i, width in enumerate([30, 20, 30, 30, 15, 60], 1):
            ws_panels.column_dimensions[get_column_letter(i)].width = width
        ws_panels.freeze_panes = "A2"

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def init_state() -> None:
    defaults = {
        "lb_xlsx_data": None,
        "lb_name": "lab-reference-data",
    }
    for key, val in defaults.items():
        if key not in st.session_state:
            st.session_state[key] = val


def reset_state() -> None:
    st.session_state.lb_xlsx_data = None
    st.session_state.lb_name = "lab-reference-data"


def render_sidebar() -> None:
    if st.button("Help & Manual", icon="📖", use_container_width=True):
        _show_manual()


def render_outputs() -> None:
    if not st.session_state.lb_xlsx_data:
        return

    st.success("Your lab reference data is ready to download.")
    col1, col2 = st.columns([3, 1])
    with col1:
        st.download_button(
            label="⬇️  Download lab reference data",
            data=st.session_state.lb_xlsx_data,
            file_name=f"{st.session_state.lb_name}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
            key="lb_dl",
        )
    with col2:
        if st.button("Start new lab data", use_container_width=True, key="lb_new"):
            reset_state()
            st.session_state.messages = []
            st.rerun()


def handle_message(user_input: str, conversation_history: str) -> None:
    with st.chat_message("assistant"):
        with st.spinner("Thinking..."):
            result = b.ProcessLabMessage(conversation_history)

        st.write(result.message)
        st.session_state.messages.append({"role": "assistant", "content": result.message})

        if result.ready_to_generate:
            try:
                lab = _build_lab_from_plan(conversation_history)

                with st.spinner("Building XLSX..."):
                    xlsx_bytes = _generate_xlsx(lab)

                st.session_state.lb_xlsx_data = xlsx_bytes
                st.session_state.lb_name = re.sub(r"[^a-z0-9-]", "-", lab.name.lower()).strip("-")
                st.rerun()
            except Exception as e:
                st.error(f"Failed to generate lab data: {e}")


# -- Private helpers -----------------------------------------------------------

def _build_lab_from_plan(conversation_history: str):
    """Build a lab object via plan + per-category test generation."""
    with st.spinner("Planning lab structure..."):
        plan = b.BuildLabPlan(conversation_history)

    all_tests = []
    for cat in plan.categories:
        total = cat.total_tests
        label = (
            f"Building {cat.name} tests ({total} test{'s' if total != 1 else ''})..."
        )
        with st.spinner(label):
            tests = b.BuildLabTestsForCategory(
                conversation_history,
                cat.code,
                cat.name,
                total,
            )

        # Tail check: catch any tests the plan undercounted
        if tests and len(tests) < total:
            with st.spinner(f"Completing {cat.name} tests..."):
                extra = b.BuildLabTestsForCategory(
                    conversation_history,
                    cat.code,
                    cat.name,
                    total - len(tests),
                )
            tests = list(tests) + list(extra)

        all_tests.extend(tests)

    return SimpleNamespace(
        name=plan.name,
        categories=plan.categories,
        tests=all_tests,
        panels=plan.panels,
        panel_test_links=plan.panel_test_links,
    )
