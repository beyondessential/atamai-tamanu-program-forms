"""
Generates a Tamanu program importer XLSX from a ProgramDefinition.

Sheet structure expected by the Tamanu program importer:
  - "Metadata" sheet: key-value program info at top, then survey table
  - One sheet per survey, named exactly after the survey name, containing questions
  - "Registry" sheet (optional): key-value registry info at top, then clinical statuses table
  - "Registry Conditions" sheet (optional): condition rows
  - "Registry Condition Categories" sheet (optional): custom category rows
"""

import re
from io import BytesIO

import openpyxl
from openpyxl.styles import Font

from baml_client.types import ProgramDefinition, ProgramRegistry, SurveyQuestion


def _flat_lower(text: str) -> str:
    """Derive a Tamanu code from a display name: lowercase, strip all non-alphanumeric chars."""
    return re.sub(r"[^a-z0-9]", "", text.lower())


def _enum_str(v) -> str:
    """Convert a SurveyType/SurveyStatus/VisibilityStatus enum to its Tamanu string.

    BAML requires PascalCase enum names (e.g. Programs, SimpleChart) but Tamanu
    expects camelCase/lowercase (e.g. programs, simpleChart). Lowercasing the first
    character handles all cases: Programs→programs, SimpleChart→simpleChart.
    QuestionType values are already correctly cased and are not passed through here.
    """
    name = v.value if hasattr(v, "value") else str(v)
    return name[0].lower() + name[1:]


# Columns for the survey question sheets, in the order the importer expects
QUESTION_COLUMNS = [
    "code",
    "type",
    "name",
    "text",
    "detail",
    "newScreen",
    "options",
    "optionLabels",
    "visibilityCriteria",
    "validationCriteria",
    "calculation",
    "config",
    "visibilityStatus",
]


def _write_header_row(ws, columns: list[str]) -> None:
    ws.append(columns)
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True)


def _build_metadata_sheet(wb: openpyxl.Workbook, program: ProgramDefinition) -> None:
    ws = wb.active
    ws.title = "Metadata"

    program_code = _flat_lower(program.program_name)

    # Key-value rows at the top (read by the importer until it hits 'code' or 'name')
    ws.append(["programCode", program_code])
    ws.append(["programName", program.program_name])
    if program.country:
        ws.append(["country", program.country])

    # Blank row for readability (the importer skips blank cells)
    ws.append([])

    # Survey table header row - must start with 'code' to trigger header detection
    survey_columns = ["code", "name", "surveyType", "status", "isSensitive", "visibilityStatus", "notifiable", "notifyEmailAddresses"]
    _write_header_row(ws, survey_columns)

    # One row per survey
    for survey in program.surveys:
        ws.append([
            _flat_lower(survey.name),
            survey.name,
            _enum_str(survey.survey_type),
            _enum_str(survey.status),
            bool(survey.is_sensitive),
            _enum_str(survey.visibility_status) if survey.visibility_status else "current",
            bool(survey.notifiable),
            survey.notify_email_addresses or "",
        ])

    # Auto-size columns
    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=0)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 50)


def _question_to_row(q: SurveyQuestion) -> list:
    return [
        q.code,
        q.type.value if hasattr(q.type, "value") else q.type,  # QuestionType stays PascalCase
        q.name,
        q.text,
        q.detail or "",
        "yes" if q.new_screen else "",
        q.options or "",
        q.option_labels or "",
        q.visibility_criteria or "",
        q.validation_criteria or "",
        q.calculation or "",
        q.config or "",
        _enum_str(q.visibility_status) if q.visibility_status else "current",
    ]


def _build_survey_sheet(wb: openpyxl.Workbook, survey_name: str, questions: list[SurveyQuestion]) -> None:
    ws = wb.create_sheet(title=survey_name)

    # Header row
    _write_header_row(ws, QUESTION_COLUMNS)

    # Question rows
    for question in questions:
        ws.append(_question_to_row(question))

    # Auto-size columns
    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=0)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 60)


def _build_registry_sheet(wb: openpyxl.Workbook, registry: ProgramRegistry) -> None:
    ws = wb.create_sheet(title="Registry")

    registry_code = _flat_lower(registry.registry_name)

    # Key-value rows at the top (read until the importer hits 'code' or 'name')
    ws.append(["registryCode", registry_code])
    ws.append(["registryName", registry.registry_name])
    ws.append(["currentlyAtType", _enum_str(registry.currently_at_type)])
    if registry.visibility_status:
        ws.append(["visibilityStatus", _enum_str(registry.visibility_status)])

    ws.append([])

    # Clinical statuses table
    _write_header_row(ws, ["code", "name", "color", "visibilityStatus"])
    for status in registry.clinical_statuses:
        ws.append([
            f"{registry_code}-{_flat_lower(status.name)}",
            status.name,
            _enum_str(status.color),
            _enum_str(status.visibility_status) if status.visibility_status else "current",
        ])

    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=0)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 50)


def _build_registry_conditions_sheet(wb: openpyxl.Workbook, registry: ProgramRegistry) -> None:
    registry_code = _flat_lower(registry.registry_name)
    ws = wb.create_sheet(title="Registry Conditions")
    _write_header_row(ws, ["code", "name", "visibilityStatus"])
    for condition in registry.conditions:
        ws.append([
            f"{registry_code}-{_flat_lower(condition.name)}",
            condition.name,
            _enum_str(condition.visibility_status) if condition.visibility_status else "current",
        ])

    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=0)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 50)


def _build_registry_condition_categories_sheet(wb: openpyxl.Workbook, registry: ProgramRegistry) -> None:
    ws = wb.create_sheet(title="Registry Condition Categories")
    _write_header_row(ws, ["code", "name", "visibilityStatus"])
    for category in registry.condition_categories:
        ws.append([
            _flat_lower(category.name),
            category.name,
            _enum_str(category.visibility_status) if category.visibility_status else "current",
        ])

    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=0)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 50)


def _to_excel_bool(value) -> bool:
    """Normalise a value (bool, 'true', 'false', 'yes', etc.) to a Python bool for Excel."""
    if isinstance(value, bool):
        return value
    return str(value).lower() in ("true", "yes", "1")


def generate_xlsx_from_program_data(program_data: dict) -> bytes:
    """
    Convert a program_data dict (from xlsx_parser or plan+batch generation)
    into XLSX bytes ready for Tamanu import.

    This is the dict-based counterpart to generate_xlsx(). It is used for:
    - Exporting programs built via the batch generation path
    - Exporting programs modified via the delta update path
    """
    wb = openpyxl.Workbook()
    ws_meta = wb.active
    ws_meta.title = "Metadata"

    # Key-value rows at the top
    ws_meta.append(["programCode", program_data.get("program_code", "")])
    ws_meta.append(["programName", program_data.get("program_name", "")])
    if program_data.get("country"):
        ws_meta.append(["country", program_data["country"]])
    ws_meta.append([])

    # Survey table
    survey_columns = ["code", "name", "surveyType", "status", "isSensitive", "visibilityStatus", "notifiable", "notifyEmailAddresses"]
    _write_header_row(ws_meta, survey_columns)

    for survey in program_data.get("surveys", []):
        ws_meta.append([
            survey.get("code", ""),
            survey.get("name", ""),
            survey.get("surveyType", "programs"),
            survey.get("status", "draft"),
            _to_excel_bool(survey.get("isSensitive", False)),
            survey.get("visibilityStatus", "current"),
            _to_excel_bool(survey.get("notifiable", False)),
            survey.get("notifyEmailAddresses", ""),
        ])

    for col in ws_meta.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=0)
        ws_meta.column_dimensions[col[0].column_letter].width = min(max_len + 2, 50)

    # Survey question sheets
    for survey in program_data.get("surveys", []):
        ws = wb.create_sheet(title=survey.get("name", "Survey"))
        _write_header_row(ws, QUESTION_COLUMNS)
        for q in survey.get("questions", []):
            ws.append([q.get(col, "") for col in QUESTION_COLUMNS])
        for col in ws.columns:
            max_len = max((len(str(cell.value or "")) for cell in col), default=0)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 60)

    # Registry sheets
    registry = program_data.get("registry")
    if registry:
        reg_kv = registry.get("kv", {})

        ws_reg = wb.create_sheet(title="Registry")
        for key, val in reg_kv.items():
            ws_reg.append([key, val])
        ws_reg.append([])
        _write_header_row(ws_reg, ["code", "name", "color", "visibilityStatus"])
        for status in registry.get("statuses", []):
            ws_reg.append([
                status.get("code", ""),
                status.get("name", ""),
                status.get("color", ""),
                status.get("visibilityStatus", "current"),
            ])
        for col in ws_reg.columns:
            max_len = max((len(str(cell.value or "")) for cell in col), default=0)
            ws_reg.column_dimensions[col[0].column_letter].width = min(max_len + 2, 50)

        if registry.get("conditions"):
            ws_cond = wb.create_sheet(title="Registry Conditions")
            _write_header_row(ws_cond, ["code", "name", "visibilityStatus"])
            for cond in registry["conditions"]:
                ws_cond.append([
                    cond.get("code", ""),
                    cond.get("name", ""),
                    cond.get("visibilityStatus", "current"),
                ])
            for col in ws_cond.columns:
                max_len = max((len(str(cell.value or "")) for cell in col), default=0)
                ws_cond.column_dimensions[col[0].column_letter].width = min(max_len + 2, 50)

        if registry.get("condition_categories"):
            ws_cat = wb.create_sheet(title="Registry Condition Categories")
            _write_header_row(ws_cat, ["code", "name", "visibilityStatus"])
            for cat in registry["condition_categories"]:
                ws_cat.append([
                    cat.get("code", ""),
                    cat.get("name", ""),
                    cat.get("visibilityStatus", "current"),
                ])
            for col in ws_cat.columns:
                max_len = max((len(str(cell.value or "")) for cell in col), default=0)
                ws_cat.column_dimensions[col[0].column_letter].width = min(max_len + 2, 50)

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def generate_xlsx(program: ProgramDefinition) -> bytes:
    """Convert a ProgramDefinition into XLSX bytes ready for download."""
    wb = openpyxl.Workbook()

    _build_metadata_sheet(wb, program)

    # Build a lookup so we can match sheet data to metadata by survey name
    sheets_by_name = {sheet.survey_name: sheet for sheet in program.survey_sheets}

    for survey in program.surveys:
        sheet_data = sheets_by_name.get(survey.name)
        if sheet_data:
            _build_survey_sheet(wb, survey.name, sheet_data.questions)
        else:
            # Create an empty sheet so the importer doesn't fail on a missing sheet
            wb.create_sheet(title=survey.name)

    if program.registry:
        _build_registry_sheet(wb, program.registry)
        if program.registry.conditions:
            _build_registry_conditions_sheet(wb, program.registry)
        if program.registry.condition_categories:
            _build_registry_condition_categories_sheet(wb, program.registry)

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
